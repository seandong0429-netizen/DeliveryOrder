from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def export_to_excel(order_data, filepath, report_type='delivery', seller_info=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Font Styles
    font_title = Font(name='Microsoft YaHei', size=16, bold=True)
    font_header = Font(name='Microsoft YaHei', size=11, bold=True)
    font_normal = Font(name='Microsoft YaHei', size=10)
    
    # Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # 1. Title
    title_text = "销售汇总表" if report_type == 'summary' else "销售出货单"
    seller_name = seller_info.get('name', "广州市 XX 办公设备有限公司") if seller_info else "广州市 XX 办公设备有限公司"

    ws.merge_cells('A1:H1')
    c1 = ws['A1']
    c1.value = seller_name
    c1.font = Font(name='Microsoft YaHei', size=20, bold=True)
    c1.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    c2 = ws['A2']
    c2.value = title_text
    c2.font = font_title
    c2.alignment = Alignment(horizontal='center')
    
    # 2. Header Info
    # Row 3: Customer | Date
    # Row 4: Address | ID
    
    # Customer Name
    ws.merge_cells('A3:D3')
    ws['A3'] = f"客户名称: {order_data.get('customer', '')}"
    ws['A3'].font = font_normal
    
    # Date (Right aligned block, but excel is grid)
    ws.merge_cells('E3:H3')
    ws['E3'] = f"单据日期: {order_data.get('date', '')}"
    ws['E3'].font = font_normal
    ws['E3'].alignment = Alignment(horizontal='right')
    
    row_offset = 3
    
    if report_type == 'delivery':
        row_offset += 1
        ws.merge_cells(f'A{row_offset}:D{row_offset}')
        ws[f'A{row_offset}'] = f"客户地址: {order_data.get('address', '')}"
        
        ws.merge_cells(f'E{row_offset}:H{row_offset}')
        ws[f'E{row_offset}'] = f"单据编号: {order_data.get('order_id', '')}"
        ws[f'E{row_offset}'].alignment = Alignment(horizontal='right')

    # 3. Table Header
    headers = ["序号", "商品名称", "规格型号", "单位", "数量", "单价", "合计", "备注"]
    header_row = row_offset + 2
    
    for idx, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=idx+1, value=h)
        cell.font = font_header
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
    # 4. Items
    items = order_data.get('items', [])
    grand_total = 0
    
    current_row = header_row + 1
    for idx, item in enumerate(items):
        # 1-based index
        ws.cell(row=current_row, column=1, value=idx+1).border = border
        ws.cell(row=current_row, column=2, value=item.get('name', '')).border = border
        ws.cell(row=current_row, column=3, value=item.get('model', '')).border = border
        ws.cell(row=current_row, column=4, value=item.get('unit', '')).border = border
        
        qty = float(item.get('qty', 0)) # Ensure float/int
        # Convert to int if integer
        if qty.is_integer(): qty = int(qty)
            
        ws.cell(row=current_row, column=5, value=qty).border = border
        
        price = float(item.get('price', 0))
        ws.cell(row=current_row, column=6, value=price).border = border
        ws.cell(row=current_row, column=6).number_format = '0.00'
        
        total = qty * price
        ws.cell(row=current_row, column=7, value=total).border = border
        ws.cell(row=current_row, column=7).number_format = '0.00'
        
        ws.cell(row=current_row, column=8, value=item.get('remark', '')).border = border
        
        # Center align most cols except Name(2) and Model(3) maybe?
        # Let's align all center or left? Usually numbers right.
        # Simple for now: Center all except Name.
        
        grand_total += total
        current_row += 1
        
    # 5. Total Row
    total_row = current_row
    ws.merge_cells(f'A{total_row}:F{total_row}')
    cell = ws[f'A{total_row}']
    cell.value = "合计"
    cell.font = font_header
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    
    # Border for merged cells needs setting on all
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).border = border
        
    # Amount
    c_amount = ws.cell(row=total_row, column=7, value=grand_total)
    c_amount.font = font_header
    c_amount.border = border
    c_amount.number_format = '0.00'
    
    # Remark slot for total row
    ws.cell(row=total_row, column=8).border = border
    
    # 6. Footer
    footer_row = total_row + 2
    ws[f'A{footer_row}'] = f"制单人: {order_data.get('maker', '管理员')}"
    ws[f'E{footer_row}'] = "签收人:"
    
    # Column Widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    wb.save(filepath)
